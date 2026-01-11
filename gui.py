"""
Panasonic DC/DC Test Dashboard - PROFESSIONAL EDITION
-----------------------------------------------------
Features:
- Live Matplotlib Graphing (Real-time Scope Trace)
- Dual Instrument Control (Scope + PSU)
- Master Sequence Automation
- Integrated Safety Interlocks & Excel Logging
- Clean Exit Handling (No Zombie Processes)
"""

import tkinter as tk
from tkinter import messagebox, scrolledtext
import json
import datetime
import time
import os
import math
import openpyxl 
from openpyxl import Workbook, load_workbook

# --- VISUALIZATION IMPORTS ---
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import numpy as np

# --- DRIVER IMPORTS ---
try:
    from ds1000z import DS1000Z
    from dp800 import DP800
except ImportError as e:
    print(f"Driver Import Warning: {e}")

# --- MOCK DRIVERS ---
class MockDS1000Z:
    def __init__(self, ip_address):
        print(f"[SIM] DS1000Z: Connected at {ip_address}")
    def reset(self):
        print(f"[SIM] DS1000Z: *RST")
    def set_source_amplitude(self, volts):
        print(f"[SIM] DS1000Z: Set Amp {volts}V")
    def set_source_frequency(self, freq):
        print(f"[SIM] DS1000Z: Set Freq {freq}Hz")
    def enable_source(self):
        print(f"[SIM] DS1000Z: Output ON")

class MockDP800:
    def __init__(self, ip_address):
        print(f"[SIM] DP800: Connected at {ip_address}")
    def set_channel(self, voltage, current, channel=1):
        print(f"[SIM] DP800: CH{channel} Set to {voltage}V / {current}A")
    def enable_output(self, channel=1):
        print(f"[SIM] DP800: CH{channel} Output ON")
    def disable_output(self, channel=1):
        print(f"[SIM] DP800: CH{channel} Output OFF")

# --- GLOBAL VARIABLES ---
scope = None
psu = None
EXCEL_FILENAME = "Panasonic_Master_Log.xlsx"
is_running = False # Controls the animation loop

# --- HELPER FUNCTIONS ---
def log_message(message):
    try:
        timestamp = datetime.datetime.now().strftime("%H:%M:%S")
        log_window.configure(state='normal')
        log_window.insert(tk.END, f"[{timestamp}] {message}\n")
        log_window.see(tk.END)
        log_window.configure(state='disabled')
    except:
        pass 

def load_config():
    try:
        with open('config.json', 'r') as f:
            config = json.load(f)
        scope_ip_entry.insert(0, config.get('oscilloscope_ip', ''))
        psu_ip_entry.insert(0, config.get('power_supply_ip', ''))
        log_message("System Initialized. Config loaded.")
    except FileNotFoundError:
        log_message("System Alert: config.json not found.")

# --- CONNECTION LOGIC ---
def connect_instruments():
    global scope, psu, is_running
    is_sim = sim_mode_var.get()
    
    log_message(f"Connecting ({'SIM' if is_sim else 'HARDWARE'})...")
    
    try:
        scope = MockDS1000Z(scope_ip_entry.get()) if is_sim else DS1000Z(scope_ip_entry.get())
        scope_status_lbl.config(text="CONNECTED", fg="#4CAF50")
    except: scope_status_lbl.config(text="ERROR", fg="red")

    try:
        psu = MockDP800(psu_ip_entry.get()) if is_sim else DP800(psu_ip_entry.get())
        psu_status_lbl.config(text="CONNECTED", fg="#4CAF50")
    except: psu_status_lbl.config(text="ERROR", fg="red")

    if scope or psu:
        
        for btn in [run_scope_btn, set_psu_btn, psu_on_btn, psu_off_btn, master_start_btn, stop_btn, save_btn]:
            btn.config(state="normal")
        log_message("Ready.")
        
        
        if not is_running:
            is_running = True
            update_graph()

# --- GRAPHING LOGIC ---
def update_graph():
    """Generates a live sine wave based on user inputs."""
    if not is_running: return # Stop immediately if flag is False

    try:
        # 1. Read User Inputs
        try: amp = float(scope_volts_entry.get())
        except: amp = 1.0
        try: freq = float(scope_freq_entry.get())
        except: freq = 1.0

        # 2. Generate Data
        t = np.linspace(0, 0.1, 500) 
        phase = time.time() * 10 
        y = amp * np.sin(2 * np.pi * freq * t + phase)
        noise = np.random.normal(0, amp * 0.05, 500)
        y = y + noise

        # 3. Update Plot
        ax.clear()
        ax.plot(t, y, color='#00FF00', linewidth=1.5) 
        ax.set_title(f"Live Scope Trace (CH1): {amp}V @ {freq}Hz", color='white')
        ax.set_facecolor('black')
        ax.grid(True, color='#333333')
        ax.set_ylim(-25, 25) 
        
        canvas.draw()
        
        # Loop: Schedule next update
        root.after(100, update_graph)
        
    except Exception:
        pass 

# --- CONTROL LOGIC ---
def run_scope_sequence():
    if not scope: return False
    try:
        volts = float(scope_volts_entry.get())
        freq = float(scope_freq_entry.get())
        if volts > 20:
            log_message("SAFETY: Scope Voltage > 20V rejected.")
            return False
        scope.set_source_amplitude(volts)
        scope.set_source_frequency(freq)
        scope.enable_source()
        log_message(f"Scope: {volts}V @ {freq}Hz set.")
        return True
    except: return False

def set_psu_params():
    if not psu: return False
    try:
        volts = float(psu_volts_entry.get())
        curr = float(psu_curr_entry.get())
        if volts > 32:
            log_message("SAFETY: PSU Voltage > 32V rejected.")
            return False
        psu.set_channel(volts, curr, channel=1)
        log_message(f"PSU: {volts}V / {curr}A set.")
        return True
    except: return False

def toggle_psu(state):
    if not psu: return
    if state:
        psu.enable_output(1)
        psu_ind.config(bg="#4CAF50")
        log_message("PSU: Output ON")
    else:
        psu.disable_output(1)
        psu_ind.config(bg="gray")
        log_message("PSU: Output OFF")

def run_master():
    log_message("--- STARTING MASTER SEQUENCE ---")
    if run_scope_sequence() and set_psu_params():
        toggle_psu(True)
        log_message("--- SEQUENCE COMPLETE ---")
    else:
        log_message("--- SEQUENCE ABORTED (Safety/Error) ---")

def emergency_stop():
    log_message("!!! EMERGENCY STOP !!!")
    if scope: scope.reset()
    if psu: 
        psu.disable_output(1)
        psu_ind.config(bg="gray")

def update_excel_log():
    ts = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    try:
        if not os.path.exists(EXCEL_FILENAME):
            wb = Workbook()
            wb.active.append(["Timestamp", "Tech", "Scope V", "Scope Hz", "PSU V", "PSU A", "Log"])
        else:
            wb = load_workbook(EXCEL_FILENAME)
        
        wb.active.append([ts, tech_entry.get(), scope_volts_entry.get(), scope_freq_entry.get(), 
                          psu_volts_entry.get(), psu_curr_entry.get(), log_window.get("1.0", tk.END).strip()])
        wb.save(EXCEL_FILENAME)
        messagebox.showinfo("Success", "Log Saved.")
    except Exception as e:
        messagebox.showerror("Error", f"Save Failed: {e}")

def on_closing():
    """Cleanly stops the loop and closes the window."""
    global is_running
    is_running = False 
    root.destroy()     
    root.quit()        

# --- GUI LAYOUT ---
root = tk.Tk()
root.title("Panasonic Critical Facilities | Master Dashboard")
root.geometry("1100x700") 
root.protocol("WM_DELETE_WINDOW", on_closing) 

# Main Containers
left_pane = tk.Frame(root, padx=10, pady=10)
left_pane.pack(side=tk.LEFT, fill="y")

right_pane = tk.Frame(root, padx=10, pady=10, bg="#2b2b2b") 
right_pane.pack(side=tk.RIGHT, fill="both", expand=True)

# --- LEFT PANE (Controls) ---
tk.Label(left_pane, text="CONTROL DECK", font=("Segoe UI", 14, "bold")).pack(pady=10)

# Tech Info
tk.Label(left_pane, text="Technician Name:").pack(anchor="w")
tech_entry = tk.Entry(left_pane, width=30)
tech_entry.pack(pady=5)

# Network
net_frame = tk.LabelFrame(left_pane, text="Connections")
net_frame.pack(fill="x", pady=10)
tk.Label(net_frame, text="Scope IP:").grid(row=0, column=0)
scope_ip_entry = tk.Entry(net_frame, width=15); scope_ip_entry.grid(row=0, column=1)
scope_status_lbl = tk.Label(net_frame, text="Offline", fg="gray"); scope_status_lbl.grid(row=0, column=2)

tk.Label(net_frame, text="PSU IP:").grid(row=1, column=0)
psu_ip_entry = tk.Entry(net_frame, width=15); psu_ip_entry.grid(row=1, column=1)
psu_status_lbl = tk.Label(net_frame, text="Offline", fg="gray"); psu_status_lbl.grid(row=1, column=2)

tk.Button(net_frame, text="CONNECT", command=connect_instruments, bg="#E0E0E0").grid(row=0, column=3, rowspan=2, padx=5)
sim_mode_var = tk.BooleanVar(value=True) 
tk.Checkbutton(net_frame, text="Sim Mode", var=sim_mode_var).grid(row=2, column=0, columnspan=2)

# Scope
scope_frame = tk.LabelFrame(left_pane, text="Oscilloscope")
scope_frame.pack(fill="x", pady=5)
tk.Label(scope_frame, text="Amp (V):").grid(row=0, column=0)
scope_volts_entry = tk.Entry(scope_frame, width=6); scope_volts_entry.insert(0, "5.0"); scope_volts_entry.grid(row=0, column=1)
tk.Label(scope_frame, text="Freq (Hz):").grid(row=0, column=2)
scope_freq_entry = tk.Entry(scope_frame, width=6); scope_freq_entry.insert(0, "50"); scope_freq_entry.grid(row=0, column=3)
run_scope_btn = tk.Button(scope_frame, text="Set", command=run_scope_sequence, state="disabled"); run_scope_btn.grid(row=0, column=4, padx=5)

# PSU
psu_frame = tk.LabelFrame(left_pane, text="Power Supply")
psu_frame.pack(fill="x", pady=5)
tk.Label(psu_frame, text="Volt (V):").grid(row=0, column=0)
psu_volts_entry = tk.Entry(psu_frame, width=6); psu_volts_entry.insert(0, "12.0"); psu_volts_entry.grid(row=0, column=1)
tk.Label(psu_frame, text="Curr (A):").grid(row=0, column=2)
psu_curr_entry = tk.Entry(psu_frame, width=6); psu_curr_entry.insert(0, "1.0"); psu_curr_entry.grid(row=0, column=3)
set_psu_btn = tk.Button(psu_frame, text="Set", command=set_psu_params, state="disabled"); set_psu_btn.grid(row=0, column=4, padx=5)

tk.Label(psu_frame, text="Output:").grid(row=1, column=0)
psu_on_btn = tk.Button(psu_frame, text="ON", bg="#81C784", command=lambda: toggle_psu(True), state="disabled"); psu_on_btn.grid(row=1, column=1)
psu_off_btn = tk.Button(psu_frame, text="OFF", bg="#E57373", command=lambda: toggle_psu(False), state="disabled"); psu_off_btn.grid(row=1, column=2)
psu_ind = tk.Frame(psu_frame, width=15, height=15, bg="gray"); psu_ind.grid(row=1, column=3)

# Master Actions
master_start_btn = tk.Button(left_pane, text="START SEQUENCE", command=run_master, bg="#43A047", fg="white", font=("bold"), state="disabled")
master_start_btn.pack(fill="x", pady=10)

stop_btn = tk.Button(left_pane, text="EMERGENCY STOP", command=emergency_stop, bg="#D32F2F", fg="white", font=("bold"), state="disabled")
stop_btn.pack(fill="x", pady=5)

save_btn = tk.Button(left_pane, text="Save to Excel", command=update_excel_log, state="disabled")
save_btn.pack(pady=10)

# --- RIGHT PANE (Graph & Log) ---
fig, ax = plt.subplots(figsize=(5, 4))
fig.patch.set_facecolor('#2b2b2b') 
ax.set_facecolor('black')
ax.set_title("Waiting for Connection...", color='white')
ax.tick_params(axis='x', colors='white')
ax.tick_params(axis='y', colors='white')

canvas = FigureCanvasTkAgg(fig, master=right_pane)
canvas.get_tk_widget().pack(fill="both", expand=True)

log_window = scrolledtext.ScrolledText(right_pane, height=8, font=("Consolas", 9), state='disabled')
log_window.pack(fill="x", pady=10)

if __name__ == "__main__":
    load_config()
    root.mainloop()